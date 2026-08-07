"""
openpyxl을 사용해 Excel(.xlsx) 문서를 생성하는 예시 스크립트.
헤더 서식, 데이터 입력, 열 너비 조정 등 기본 기능에 더해
ID 열 자동 추가, 제목/인용구/본문 폰트 스타일링, 비교 데이터 차트 자동 생성을 포함한
샘플 워크북을 만든다.
"""

import copy

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference


# ---------------------------------------------------------------------------
# 제목 / 인용구 / 차트 판단에 사용하는 키워드
# ---------------------------------------------------------------------------
TITLE_KEYWORDS = ("제목", "title", "headline", "heading")
QUOTE_KEYWORDS = ("인용구", "quote", "quotation", "citation")
QUOTE_CHARS = ('"', "'", "“", "”", "‘", "’", "「", "」")
TIME_KEYWORDS = ("연도", "년도", "월", "날짜", "일자", "date", "year", "month")
CHART_KEYWORDS = (
    "성장률", "증감률", "변화율", "증가율", "감소율", "전년", "전월", "비교", "비율",
    "실적", "수치", "매출", "사용자", "점수",
    "percentage", "growth", "rate", "change", "comparison", "revenue", "value",
)


def add_id_column(ws, header_row: int = 1, data_start_row: int = 2, id_header: str = "ID") -> None:
    """A열에 ID 열을 삽입하고 데이터가 있는 행에만 1, 2, 3... 순번을 채운다."""
    ws.insert_cols(1)

    header_cell = ws.cell(row=header_row, column=1, value=id_header)
    ref_header_cell = ws.cell(row=header_row, column=2)
    header_cell.font = copy.copy(ref_header_cell.font)
    header_cell.fill = copy.copy(ref_header_cell.fill)
    header_cell.alignment = copy.copy(ref_header_cell.alignment)
    header_cell.border = copy.copy(ref_header_cell.border)

    current_id = 1
    for row_idx in range(data_start_row, ws.max_row + 1):
        row_has_data = any(
            ws.cell(row=row_idx, column=col_idx).value not in (None, "")
            for col_idx in range(2, ws.max_column + 1)
        )
        if not row_has_data:
            continue

        ref_cell = ws.cell(row=row_idx, column=2)
        id_cell = ws.cell(row=row_idx, column=1, value=current_id)
        id_cell.font = copy.copy(ref_cell.font)
        id_cell.alignment = copy.copy(ref_cell.alignment)
        id_cell.border = copy.copy(ref_cell.border)
        current_id += 1


def _classify_columns(ws, header_row: int = 1):
    """헤더 텍스트를 보고 제목 열 / 인용구 열의 인덱스를 판별한다."""
    title_cols, quote_cols = set(), set()
    for col_idx in range(1, ws.max_column + 1):
        header_value = ws.cell(row=header_row, column=col_idx).value
        header_text = str(header_value).lower() if header_value is not None else ""
        if any(keyword.lower() in header_text for keyword in TITLE_KEYWORDS):
            title_cols.add(col_idx)
        if any(keyword.lower() in header_text for keyword in QUOTE_KEYWORDS):
            quote_cols.add(col_idx)
    return title_cols, quote_cols


def _looks_like_quote(value) -> bool:
    if not isinstance(value, str) or len(value) < 2:
        return False
    return value[0] in QUOTE_CHARS and value[-1] in QUOTE_CHARS


def apply_cell_styles(ws, header_row: int = 1, data_start_row: int = 2, body_size: int = 10) -> None:
    """제목 열은 굵게, 인용구 열(혹은 인용부호로 감싸진 값)은 기울임, 나머지는 10pt 본문으로 처리한다.

    기존 폰트의 이름/색상/밑줄 등은 유지하고 크기와 굵기/기울임만 갱신한다.
    """
    title_cols, quote_cols = _classify_columns(ws, header_row)

    for row_idx in range(data_start_row, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is None:
                continue

            base_font = copy.copy(cell.font)
            is_bold = bool(base_font.bold) or (col_idx in title_cols)
            is_italic = bool(base_font.italic) or (col_idx in quote_cols) or _looks_like_quote(cell.value)

            cell.font = Font(
                name=base_font.name,
                size=body_size,
                bold=is_bold,
                italic=is_italic,
                color=base_font.color,
                underline=base_font.underline,
                strike=base_font.strike,
                vertAlign=base_font.vertAlign,
            )


def detect_numeric_columns(ws, header_row: int = 1, data_start_row: int = 2):
    """차트 키워드에 해당하는 헤더를 가지면서 숫자 데이터가 2개 이상인 열을 찾는다."""
    candidates = []
    for col_idx in range(1, ws.max_column + 1):
        header_value = ws.cell(row=header_row, column=col_idx).value
        if header_value is None:
            continue

        header_text = str(header_value)
        header_lower = header_text.lower()
        if not any(keyword.lower() in header_lower for keyword in CHART_KEYWORDS):
            continue

        numeric_values = [
            cell.value
            for cell in (
                ws.cell(row=row_idx, column=col_idx)
                for row_idx in range(data_start_row, ws.max_row + 1)
            )
            if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool)
        ]
        if len(numeric_values) >= 2:
            candidates.append({"col_idx": col_idx, "header": header_text, "count": len(numeric_values)})

    return candidates


def detect_chart_type(ws, columns, header_row: int = 1) -> str:
    """시간 흐름에 따른 데이터면 line, 범주별 비교 데이터면 bar를 반환한다."""
    for col_idx in range(1, ws.max_column + 1):
        header_value = ws.cell(row=header_row, column=col_idx).value
        if header_value is None:
            continue
        if any(keyword.lower() in str(header_value).lower() for keyword in TIME_KEYWORDS):
            return "line"
    return "bar"


def create_data_charts(ws, header_row: int = 1, data_start_row: int = 2) -> None:
    """비교 가능한 숫자 데이터가 있으면 표와 겹치지 않게 오른쪽에 차트를 생성한다."""
    numeric_cols = detect_numeric_columns(ws, header_row, data_start_row)
    if not numeric_cols or numeric_cols[0]["count"] < 2:
        return

    chart_kind = detect_chart_type(ws, numeric_cols, header_row)
    chart = LineChart() if chart_kind == "line" else BarChart()
    chart.title = " / ".join(col["header"] for col in numeric_cols)
    chart.y_axis.title = "값"

    label_col = 2 if ws.max_column >= 2 else 1
    label_header = ws.cell(row=header_row, column=label_col).value
    chart.x_axis.title = str(label_header) if label_header else "항목"
    chart.style = 10
    chart.legend.position = "b"

    min_col = min(col["col_idx"] for col in numeric_cols)
    max_col = max(col["col_idx"] for col in numeric_cols)
    max_row = ws.max_row

    data = Reference(ws, min_col=min_col, max_col=max_col, min_row=header_row, max_row=max_row)
    categories = Reference(ws, min_col=label_col, max_col=label_col, min_row=data_start_row, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    anchor_col = get_column_letter(ws.max_column + 2)
    ws.add_chart(chart, f"{anchor_col}2")


def autofit_columns(ws, padding: int = 6) -> None:
    """모든 열의 내용 길이에 맞춰 너비를 자동 조정한다."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = max((len(str(cell.value)) for cell in col_cells if cell.value is not None), default=0)
        ws.column_dimensions[col_letter].width = max_len + padding


def apply_workbook_formatting(wb) -> None:
    """ID 열 추가, 폰트 스타일링, 열 너비 조정, 차트 생성을 각 시트에 순서대로 적용한다.

    개별 단계가 실패하더라도 나머지 시트/워크북 저장에는 영향이 없도록 예외를 흡수한다.
    """
    for ws in wb.worksheets:
        try:
            add_id_column(ws)
        except Exception as exc:
            print(f"[경고] '{ws.title}' 시트의 ID 열 추가에 실패했습니다: {exc}")

        try:
            apply_cell_styles(ws)
        except Exception as exc:
            print(f"[경고] '{ws.title}' 시트의 폰트 스타일 적용에 실패했습니다: {exc}")

        try:
            autofit_columns(ws)
        except Exception as exc:
            print(f"[경고] '{ws.title}' 시트의 열 너비 조정에 실패했습니다: {exc}")

        try:
            create_data_charts(ws)
        except Exception as exc:
            print(f"[경고] '{ws.title}' 시트의 차트 생성에 실패했습니다: {exc}")

        ws.freeze_panes = "A2"


def create_sample_workbook(output_path: str = "output.xlsx") -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "직원 목록"

    headers = ["이름", "부서", "직급", "입사일"]
    rows_data = [
        ("홍길동", "개발팀", "대리", "2022-03-02"),
        ("김철수", "기획팀", "과장", "2019-07-15"),
        ("이영희", "디자인팀", "사원", "2023-01-09"),
    ]

    header_fill = PatternFill(start_color="2E74B5", end_color="2E74B5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # 헤더 행
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # 데이터 행
    for row_idx, row in enumerate(rows_data, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    # ID 열 추가 / 제목·인용구·본문 폰트 스타일링 / 열 너비 조정 / 차트 생성 / 첫 행 고정
    apply_workbook_formatting(wb)

    wb.save(output_path)
    print(f"문서가 저장되었습니다: {output_path}")


if __name__ == "__main__":
    create_sample_workbook()
