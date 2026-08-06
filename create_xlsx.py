"""
openpyxl을 사용해 Excel(.xlsx) 문서를 생성하는 예시 스크립트.
헤더 서식, 데이터 입력, 열 너비 조정 등 기본 기능을 포함한 샘플 워크북을 만든다.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_sample_workbook(output_path: str = "output.xlsx") -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "직원 목록"

    headers = ["이름", "부서", "직급", "입사일"]
    rows_data = [
        ("홍길동", "개발팀", "대리", "2022-03-02"),
        ("김철수", "기획팀", "과장", "2019-07-15"),
        ("이영희", "디자인팀", "사원", "2023-01-09"),
    ]

    header_fill = PatternFill(start_color="2E74B5", end_color="2E74B5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # 헤더 행
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # 데이터 행
    for row_idx, row in enumerate(rows_data, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    # 열 너비 자동 조정
    for col_idx, title in enumerate(headers, start=1):
        max_len = max(
            [len(str(title))] + [len(str(row[col_idx - 1])) for row in rows_data]
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 6

    # 첫 행 고정
    ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f"문서가 저장되었습니다: {output_path}")


if __name__ == "__main__":
    create_sample_workbook()
