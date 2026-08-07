#!/usr/bin/env python3
"""
openpyxl 기반 범용 Excel(.xlsx) 생성 스크립트.

JSON으로 표 데이터(헤더/행)를 받아, 시트마다
- ID 열 자동 추가
- 제목/인용구/본문 폰트 스타일링 (10pt 본문, 제목 굵게, 인용구 기울임)
- 성장률/비교 등 숫자 데이터에 대한 차트 자동 생성
- 열 너비 자동 조정, 첫 행 고정
을 적용한 워크북을 만든다.

원본은 프로젝트 루트의 create_xlsx.py이며, 이 스크립트는 그 로직을
특정 샘플 데이터가 아닌 임의의 JSON 입력에 대해 동작하도록 일반화한 버전이다.

사용법:
    python generate_xlsx.py --input content.json --output report.xlsx [--no-id] [--no-charts]

입력 JSON 스키마는 references/content_schema.md 참고.
"""

import argparse
import copy
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference


# ---------------------------------------------------------------------------
# 제목 / 인용구 / 차트 판단에 사용하는 키워드
# ---------------------------------------------------------------------------
TITLE_KEYWORDS = ("제목", "title", "headline", "heading")
QUOTE_KEYWORDS = ("인용구", "quote", "quotation", "citation")
QUOTE_CHARS = ('"', "'", "“", "”", "‘", "’", "「", "」")
TIME_KEYWORDS = ("연도", "년도", "월", "날짜", "일자", "date", "year", "month")
CHART_KEYWORDS = (
    "성장률", "증감률", "변화율", "증가율", "감소율", "전년", "전월", "비교", "비율",
    "실적", "수치", "매출", "사용자", "점수",
    "percentage", "growth", "rate", "change", "comparison", "revenue", "value",
)

DEFAULT_HEADER_COLOR = "2E74B5"


# ---------------------------------------------------------------------------
# 시트 기본 골격 (헤더 + 데이터 행) 작성
# ---------------------------------------------------------------------------
def write_table(ws, headers, rows, header_color: str = DEFAULT_HEADER_COLOR) -> None:
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")


# ---------------------------------------------------------------------------
# 1) ID 열 추가
# ---------------------------------------------------------------------------
def add_id_column(ws, header_row: int = 1, data_start_row: int = 2, id_header: str = "ID") -> None:
    """A열에 ID 열을 삽입하고 데이터가 있는 행에만 1, 2, 3... 순번을 채운다."""
    ws.insert_cols(1)

    header_cell = ws.cell(row=header_row, column=1, value=id_header)
    ref_header_cell = ws.cell(row=header_row, column=2)
    header_cell.font = copy.copy(ref_header_cell.font)
    header_cell.fill = copy.copy(ref_header_cell.fill)
    header_cell.alignment = copy.copy(ref_header_cell.alignment)
    header_cell.border = copy.copy(ref_header_cell.border)

    current_id = 1
    for row_idx in range(data_start_row, ws.max_row + 1):
        row_has_data = any(
            ws.cell(row=row_idx, column=col_idx).value not in (None, "")
            for col_idx in range(2, ws.max_column + 1)
        )
        if not row_has_data:
            continue

        ref_cell = ws.cell(row=row_idx, column=2)
        id_cell = ws.cell(row=row_idx, column=1, value=current_id)
        id_cell.font = copy.copy(ref_cell.font)
        id_cell.alignment = copy.copy(ref_cell.alignment)
        id_cell.border = copy.copy(ref_cell.border)
        current_id += 1


# ---------------------------------------------------------------------------
# 2) 제목 / 인용구 / 본문 폰트 스타일링
# ---------------------------------------------------------------------------
def _classify_columns(ws, header_row: int = 1, extra_title_headers=(), extra_quote_headers=()):
    """헤더 텍스트를 보고 제목 열 / 인용구 열의 인덱스를 판별한다.

    extra_title_headers / extra_quote_headers 로 자동 키워드 판별 대신
    특정 헤더명을 명시적으로 지정할 수도 있다.
    """
    title_cols, quote_cols = set(), set()
    extra_title_lower = {h.lower() for h in extra_title_headers}
    extra_quote_lower = {h.lower() for h in extra_quote_headers}

    for col_idx in range(1, ws.max_column + 1):
        header_value = ws.cell(row=header_row, column=col_idx).value
        header_text = str(header_value).lower() if header_value is not None else ""

        if header_text in extra_title_lower or any(k.lower() in header_text for k in TITLE_KEYWORDS):
            title_cols.add(col_idx)
        if header_text in extra_quote_lower or any(k.lower() in header_text for k in QUOTE_KEYWORDS):
            quote_cols.add(col_idx)

    return title_cols, quote_cols


def _looks_like_quote(value) -> bool:
    if not isinstance(value, str) or len(value) < 2:
        return False
    return value[0] in QUOTE_CHARS and value[-1] in QUOTE_CHARS


def apply_cell_styles(
    ws,
    header_row: int = 1,
    data_start_row: int = 2,
    body_size: int = 10,
    title_headers=(),
    quote_headers=(),
) -> None:
    """제목 열은 굵게, 인용구 열(혹은 인용부호로 감싸진 값)은 기울임, 나머지는 body_size pt 본문으로 처리한다.

    기존 폰트의 이름/색상/밑줄 등은 유지하고 크기와 굵기/기울임만 갱신한다.
    """
    title_cols, quote_cols = _classify_columns(ws, header_row, title_headers, quote_headers)

    for row_idx in range(data_start_row, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is None:
                continue

            base_font = copy.copy(cell.font)
            is_bold = bool(base_font.bold) or (col_idx in title_cols)
            is_italic = bool(base_font.italic) or (col_idx in quote_cols) or _looks_like_quote(cell.value)

            cell.font = Font(
                name=base_font.name,
                size=body_size,
                bold=is_bold,
                italic=is_italic,
                color=base_font.color,
                underline=base_font.underline,
                strike=base_font.strike,
                vertAlign=base_font.vertAlign,
            )


# ---------------------------------------------------------------------------
# 3) 차트 대상 숫자 열 탐지 및 차트 생성
# ---------------------------------------------------------------------------
def detect_numeric_columns(ws, header_row: int = 1, data_start_row: int = 2):
    """차트 키워드에 해당하는 헤더를 가지면서 숫자 데이터가 2개 이상인 열을 찾는다."""
    candidates = []
    for col_idx in range(1, ws.max_column + 1):
        header_value = ws.cell(row=header_row, column=col_idx).value
        if header_value is None:
            continue

        header_text = str(header_value)
        header_lower = header_text.lower()
        if not any(keyword.lower() in header_lower for keyword in CHART_KEYWORDS):
            continue

        numeric_values = [
            cell.value
            for cell in (
                ws.cell(row=row_idx, column=col_idx)
                for row_idx in range(data_start_row, ws.max_row + 1)
            )
            if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool)
        ]
        if len(numeric_values) >= 2:
            candidates.append({"col_idx": col_idx, "header": header_text, "count": len(numeric_values)})

    return candidates


def detect_chart_type(ws, columns, header_row: int = 1) -> str:
    """시간 흐름에 따른 데이터면 line, 범주별 비교 데이터면 bar를 반환한다."""
    for col_idx in range(1, ws.max_column + 1):
        header_value = ws.cell(row=header_row, column=col_idx).value
        if header_value is None:
            continue
        if any(keyword.lower() in str(header_value).lower() for keyword in TIME_KEYWORDS):
            return "line"
    return "bar"


def create_data_charts(ws, header_row: int = 1, data_start_row: int = 2) -> None:
    """비교 가능한 숫자 데이터가 있으면 표와 겹치지 않게 오른쪽에 차트를 생성한다."""
    numeric_cols = detect_numeric_columns(ws, header_row, data_start_row)
    if not numeric_cols or numeric_cols[0]["count"] < 2:
        return

    chart_kind = detect_chart_type(ws, numeric_cols, header_row)
    chart = LineChart() if chart_kind == "line" else BarChart()
    chart.title = " / ".join(col["header"] for col in numeric_cols)
    chart.y_axis.title = "값"

    label_col = 2 if ws.max_column >= 2 else 1
    label_header = ws.cell(row=header_row, column=label_col).value
    chart.x_axis.title = str(label_header) if label_header else "항목"
    chart.style = 10
    chart.legend.position = "b"

    min_col = min(col["col_idx"] for col in numeric_cols)
    max_col = max(col["col_idx"] for col in numeric_cols)
    max_row = ws.max_row

    data = Reference(ws, min_col=min_col, max_col=max_col, min_row=header_row, max_row=max_row)
    categories = Reference(ws, min_col=label_col, max_col=label_col, min_row=data_start_row, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    anchor_col = get_column_letter(ws.max_column + 2)
    ws.add_chart(chart, f"{anchor_col}2")


# ---------------------------------------------------------------------------
# 4) 열 너비 자동 조정
# ---------------------------------------------------------------------------
def autofit_columns(ws, padding: int = 6) -> None:
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = max((len(str(cell.value)) for cell in col_cells if cell.value is not None), default=0)
        ws.column_dimensions[col_letter].width = max_len + padding


# ---------------------------------------------------------------------------
# 5) 시트/워크북 단위 오케스트레이션
# ---------------------------------------------------------------------------
def apply_sheet_formatting(
    ws,
    use_id_column: bool = True,
    use_charts: bool = True,
    id_header: str = "ID",
    title_headers=(),
    quote_headers=(),
) -> None:
    """ID 열 추가, 폰트 스타일링, 열 너비 조정, 차트 생성을 한 시트에 순서대로 적용한다.

    개별 단계가 실패하더라도 나머지 시트/워크북 저장에는 영향이 없도록 예외를 흡수한다.
    """
    if use_id_column:
        try:
            add_id_column(ws, id_header=id_header)
        except Exception as exc:
            print(f"[경고] '{ws.title}' 시트의 ID 열 추가에 실패했습니다: {exc}")

    try:
        apply_cell_styles(ws, title_headers=title_headers, quote_headers=quote_headers)
    except Exception as exc:
        print(f"[경고] '{ws.title}' 시트의 폰트 스타일 적용에 실패했습니다: {exc}")

    try:
        autofit_columns(ws)
    except Exception as exc:
        print(f"[경고] '{ws.title}' 시트의 열 너비 조정에 실패했습니다: {exc}")

    if use_charts:
        try:
            create_data_charts(ws)
        except Exception as exc:
            print(f"[경고] '{ws.title}' 시트의 차트 생성에 실패했습니다: {exc}")

    ws.freeze_panes = "A2"


def build_workbook(content: dict, use_id_column: bool = True, use_charts: bool = True) -> Workbook:
    sheets = content.get("sheets")
    if not sheets:
        raise ValueError("입력 JSON에 'sheets' 배열이 없습니다.")

    wb = Workbook()
    wb.remove(wb.active)

    for sheet_def in sheets:
        name = str(sheet_def.get("name", "Sheet1"))[:31]
        headers = sheet_def.get("headers", [])
        rows = sheet_def.get("rows", [])
        header_color = sheet_def.get("header_color", DEFAULT_HEADER_COLOR)
        sheet_use_id = sheet_def.get("add_id_column", use_id_column)
        id_header = sheet_def.get("id_header", "ID")
        title_headers = sheet_def.get("title_columns", [])
        quote_headers = sheet_def.get("quote_columns", [])

        ws = wb.create_sheet(title=name)
        write_table(ws, headers, rows, header_color=header_color)
        apply_sheet_formatting(
            ws,
            use_id_column=sheet_use_id,
            use_charts=use_charts,
            id_header=id_header,
            title_headers=title_headers,
            quote_headers=quote_headers,
        )

    return wb


def main() -> None:
    parser = argparse.ArgumentParser(description="JSON 데이터를 서식이 적용된 .xlsx 파일로 변환한다.")
    parser.add_argument("--input", required=True, help="입력 content JSON 경로")
    parser.add_argument("--output", required=True, help="출력 .xlsx 경로")
    parser.add_argument("--no-id", action="store_true", help="ID 열을 추가하지 않음")
    parser.add_argument("--no-charts", action="store_true", help="차트를 생성하지 않음")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"입력 파일을 찾을 수 없습니다: {input_path}")
        sys.exit(1)

    content = json.loads(input_path.read_text(encoding="utf-8"))

    try:
        wb = build_workbook(content, use_id_column=not args.no_id, use_charts=not args.no_charts)
    except Exception as exc:
        print(f"워크북 생성에 실패했습니다: {exc}")
        sys.exit(1)

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"문서가 저장되었습니다: {output_path}")


if __name__ == "__main__":
    main()
